import redis
import string
import random
from openpyxl import Workbook


def getdata(inputdata):
    file_name = './log/' + ''.join(
        random.sample(string.ascii_letters + string.digits, 10)) + '.xlsx'  # 我存放在log目录下 这里可以自定义
    r = redis.Redis(host='127.0.0.1', port=6379, db=6)  # 连接本地redis数据库的第6个DB
    feeds = r.keys("*" + inputdata + "*")  # 在redis的hash中 匹配用户输入的字符
    wb = Workbook()  # 创建工作簿
    ws = wb.active  # 激活工作表
    ws['A1'] = '漏洞编号'
    ws['B1'] = '披露日期'
    ws['C1'] = 'CVSS评级'
    ws['D1'] = '漏洞名称'
    ws['E1'] = '漏洞描述'
    for key in feeds:
        getdata = r.hmget(key, '编号', '漏洞详情', '等级', '漏洞名称', '漏洞描述')
        ws.append(
            [getdata[0].decode('UTF8'), getdata[1].decode('UTF8'), getdata[2].decode('UTF8'), getdata[3].decode('UTF8'),
             getdata[4].decode('UTF8')])
        wb.save(file_name)
    return 'http://www.liock.cn/wx/log/' + file_name[6:]
    # 返回路径 file_name处我用了/log/ 所以拼接的时候就从第六个字符开始的
